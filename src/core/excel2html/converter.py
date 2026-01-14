"""
Excel 转 HTML 转换器 - RAG 增强版
将 Excel 文件转换为 RAG 优化的 HTML 格式

增强功能：
1. 上下文硬编码 - 注入文件名、Sheet名等元数据
2. 幽灵标题 - 添加同义词和关键检索词
3. 表头降维 - 把父级标题拼接到子级标题（针对多层表头）
4. 合并单元格智能处理
5. 注释提取和分发
"""

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from loguru import logger

from ..models import MergedCellInfo


@dataclass
class ExcelToHtmlConverter:
    """Excel 转 HTML 转换器"""

    keywords: list[str] | None = None
    _merged_info: dict[tuple[int, int], MergedCellInfo] = field(
        default_factory=dict, init=False, repr=False
    )

    def convert(self, excel_path: Path, output_path: Path | None = None) -> Path | None:
        """执行转换"""
        source_path = Path(excel_path) if not isinstance(excel_path, Path) else excel_path

        if not source_path.exists():
            logger.error(f"找不到文件 '{source_path}'")
            return None

        out_path = self._determine_output_path(source_path, output_path)
        filename = source_path.name

        logger.info(f"正在处理: {filename}")
        self._log_features()

        try:
            workbook = openpyxl.load_workbook(str(source_path), data_only=False)
        except Exception as e:
            logger.error(f"解析失败: {e}")
            return None

        html_parts = self._convert_workbook(workbook, filename)

        return self._write_output(out_path, html_parts)

    def _determine_output_path(self, source_path: Path, output_path: Path | None) -> Path:
        """确定输出路径"""
        if output_path:
            return Path(output_path)
        return source_path.with_suffix("").with_name(f"{source_path.stem}_middle.html")

    def _log_features(self) -> None:
        """记录启用的功能"""
        features = "上下文硬编码 ✓ | 表头降维 ✓ | 合并单元格 ✓ | 注释提取 ✓"
        if self.keywords:
            features += f" | 幽灵标题 ✓ ({len(self.keywords)}个关键词)"
        else:
            features += " | 幽灵标题 ✗ (未提供关键词)"
        logger.info(f"增强功能: {features}")

    def _convert_workbook(self, workbook, filename: str) -> list[str]:
        """转换整个工作簿"""
        html_parts = []
        for sheet in workbook.worksheets:
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue
            html_parts.append(self._sheet_to_html(sheet, filename))
        return html_parts

    def _write_output(self, out_path: Path, html_parts: list[str]) -> Path | None:
        """写入输出文件"""
        try:
            out_path.write_text("\n".join(html_parts), encoding="utf-8")
            logger.info(f"转换成功！输出: {out_path.absolute()}")
            return out_path
        except OSError as e:
            logger.error(f"写入文件失败: {e}")
            return None

    def _sheet_to_html(self, sheet, filename: str) -> str:
        """将单个 sheet 转换为 RAG 增强的 HTML 表格"""
        self._merged_info = self._extract_merged_cells(sheet)
        header_rows = self._detect_header_rows(sheet)
        flattened_headers = self._build_flattened_headers(sheet, header_rows)

        footer_notes, data_end_row = self._detect_footer_notes(sheet, header_rows)
        notes_dict = self._parse_notes_with_keys(footer_notes)

        header_text = " ".join(flattened_headers.values())
        header_note_refs = self._extract_note_references(header_text)

        html_parts = self._build_html_parts(
            sheet,
            filename,
            flattened_headers,
            notes_dict,
            header_note_refs,
            header_rows,
            data_end_row,
        )

        return "\n".join(html_parts)

    def _extract_merged_cells(self, sheet) -> dict[tuple[int, int], MergedCellInfo]:
        """获取所有合并单元格的信息"""
        merged_info: dict[tuple[int, int], MergedCellInfo] = {}

        for merged_range in sheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            origin_value = sheet.cell(row=min_row, column=min_col).value

            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    is_origin = r == min_row and c == min_col
                    merged_info[(r, c)] = MergedCellInfo(
                        value=str(origin_value) if origin_value is not None else None,
                        rowspan=rowspan if is_origin else 0,
                        colspan=colspan if is_origin else 0,
                        is_origin=is_origin,
                        skip=not is_origin,
                    )

        return merged_info

    def _detect_header_rows(self, sheet, max_check_rows: int = 5) -> int:
        """检测表头行数"""
        header_rows = 1

        for row_idx in range(1, min(max_check_rows + 1, sheet.max_row + 1)):
            has_colspan = False
            for col_idx in range(1, sheet.max_column + 1):
                info = self._merged_info.get((row_idx, col_idx))
                if info and info.colspan > 1:
                    has_colspan = True
                    break

            if has_colspan:
                header_rows = max(header_rows, row_idx + 1)

        return min(header_rows, sheet.max_row)

    def _build_flattened_headers(self, sheet, header_rows: int) -> dict[int, str]:
        """构建降维后的表头"""
        if header_rows <= 1:
            return self._build_single_row_headers(sheet)

        return self._build_multi_row_headers(sheet, header_rows)

    def _build_single_row_headers(self, sheet) -> dict[int, str]:
        """构建单行表头"""
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col_idx).value
            headers[col_idx] = str(value) if value else f"列{col_idx}"
        return headers

    def _build_multi_row_headers(self, sheet, header_rows: int) -> dict[int, str]:
        """构建多行表头（降维）"""
        col_values: dict[int, list[str]] = {col: [] for col in range(1, sheet.max_column + 1)}

        for row_idx in range(1, header_rows + 1):
            for col_idx in range(1, sheet.max_column + 1):
                info = self._merged_info.get((row_idx, col_idx))
                if info:
                    value = info.value
                    col_values[col_idx].append(str(value) if value else "")
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    col_values[col_idx].append(self._format_cell_value(cell))

        headers = {}
        for col_idx, values in col_values.items():
            unique_values = []
            for v in values:
                v = v.strip()
                if v and (not unique_values or v != unique_values[-1]):
                    unique_values.append(v)
            headers[col_idx] = "-".join(unique_values) if unique_values else f"列{col_idx}"

        return headers

    def _format_cell_value(self, cell) -> str:
        """格式化单元格值"""
        value = cell.value
        if value is None:
            return ""

        number_format = cell.number_format or "General"

        if isinstance(value, datetime):
            return self._format_datetime(value, number_format)

        if not isinstance(value, (int, float)):
            return str(value)

        return self._format_number(value, number_format)

    def _format_datetime(self, value: datetime, number_format: str) -> str:
        """格式化日期时间"""
        if "H" in number_format or "h" in number_format:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")

    def _format_number(self, value: int | float, number_format: str) -> str:
        """格式化数字"""
        if "%" in number_format:
            return self._format_percentage(value, number_format)

        if "E" in number_format.upper() and number_format != "General":
            return self._format_scientific(value, number_format)

        if "#,##" in number_format or ",0" in number_format:
            return self._format_currency(value, number_format)

        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)

    def _format_percentage(self, value: float, number_format: str) -> str:
        """格式化百分比"""
        decimal_match = re.search(r"0\.(0+)%", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        return f"{value * 100:.{decimals}f}%"

    def _format_scientific(self, value: float, number_format: str) -> str:
        """格式化科学计数法"""
        decimal_match = re.search(r"0\.(0+)E", number_format, re.IGNORECASE)
        decimals = len(decimal_match.group(1)) if decimal_match else 2
        return f"{value:.{decimals}E}"

    def _format_currency(self, value: float, number_format: str) -> str:
        """格式化货币"""
        decimal_match = re.search(r"0\.(0+)", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        formatted = f"{value:,.{decimals}f}"
        if "¥" in number_format or "￥" in number_format:
            return f"¥{formatted}"
        elif "$" in number_format:
            return f"${formatted}"
        return formatted

    def _detect_footer_notes(self, sheet, header_rows: int) -> tuple[list[str], int]:
        """检测表格末尾的注释行"""
        notes: list[str] = []
        note_patterns = [
            "注",
            "备注",
            "说明",
            "注意",
            "*",
            "※",
            "●",
            "◆",
            "△",
            "▲",
            "[注",
            "（注",
            "(注",
        ]

        for row_idx in range(sheet.max_row, header_rows, -1):
            content, is_note = self._check_note_row(sheet, row_idx, note_patterns)

            if not content:
                continue

            if is_note:
                notes.insert(0, content)
            else:
                break

        data_end_row = sheet.max_row - len(notes)
        return notes, data_end_row

    def _check_note_row(self, sheet, row_idx: int, note_patterns: list[str]) -> tuple[str, bool]:
        """检查是否为注释行"""
        filled_cols = 0
        content = ""
        is_merged_wide = False

        for col_idx in range(1, sheet.max_column + 1):
            info = self._merged_info.get((row_idx, col_idx))
            cell = sheet.cell(row=row_idx, column=col_idx)

            if info and info.is_origin and info.colspan > sheet.max_column // 2:
                is_merged_wide = True
                content = info.value if info.value else ""
                break

            if info and info.skip:
                continue

            if cell.value:
                filled_cols += 1
                if not content:
                    content = str(cell.value)

        content = content.strip()
        if not content:
            return "", False

        is_note = is_merged_wide or (
            filled_cols <= 2 and any(content.startswith(p) for p in note_patterns)
        )

        return content, is_note

    def _parse_notes_with_keys(self, notes_list: list[str]) -> dict[str, str]:
        """解析注释列表，提取注释编号和内容"""
        notes_dict: dict[str, str] = {}

        for note in notes_list:
            note = note.strip()
            if not note:
                continue

            split_pattern = r"(?=\[注[\d、,，]+\]|\[备注\d*\]|\[说明\d*\])"
            parts = re.split(split_pattern, note)
            parts = [p.strip() for p in parts if p.strip()]

            if len(parts) > 1:
                for part in parts:
                    self._parse_single_note(part, notes_dict)
            else:
                self._parse_single_note(note, notes_dict)

        return notes_dict

    def _parse_single_note(self, note: str, notes_dict: dict[str, str]) -> None:
        """解析单个注释"""
        note = note.strip()
        if not note:
            return

        # 匹配 [注3、4、5] 格式
        multi_num_match = re.match(r"^\[(注)([\d、,，]+)\]", note)
        if multi_num_match:
            prefix = multi_num_match.group(1)
            nums_str = multi_num_match.group(2)
            nums = re.split(r"[、,，]", nums_str)
            for num in nums:
                num = num.strip()
                if num:
                    notes_dict[f"{prefix}{num}"] = note
            return

        # 匹配 [注1] 格式
        bracket_match = re.match(r"^\[([注备说][注明意]?\d*)\]", note)
        if bracket_match:
            notes_dict[bracket_match.group(1)] = note
            return

        # 匹配 （注1） 或 (注1) 格式
        paren_match = re.match(r"^[（\(]([注备说][注明意]?\d*)[）\)]", note)
        if paren_match:
            notes_dict[paren_match.group(1)] = note
            return

        # 匹配 注1：内容 格式
        plain_match = re.match(r"^(注\d*|备注\d*|说明\d*|注意\d*)[：:．.、]?\s*", note)
        if plain_match:
            key_match = re.match(r"^(注\d*|备注\d*|说明\d*|注意\d*)", note)
            if key_match:
                notes_dict[key_match.group(1)] = note
            return

        # 匹配特殊符号开头
        if note[0] in "*※●◆△▲":
            notes_dict[note[0]] = note
            return

        # 无法识别的注释
        notes_dict[note[:10]] = note

    def _extract_note_references(self, text: str) -> set[str]:
        """从文本中提取注释引用"""
        refs: set[str] = set()

        # 匹配合并格式: [注1、2、3]
        multi_refs = re.findall(r"\[(注)([\d、,，]+)\]", text)
        for prefix, nums_str in multi_refs:
            nums = re.split(r"[、,，]", nums_str)
            for num in nums:
                num = num.strip()
                if num:
                    refs.add(f"{prefix}{num}")

        # 匹配单个方括号注释
        bracket_refs = re.findall(r"\[(注\d*|备注\d*|说明\d*|注意\d*)\]", text)
        refs.update(bracket_refs)

        # 匹配无方括号的注释引用
        superscript_refs = re.findall(r"[^\[](注\d+)(?:[：:）\)]|$|\s)", text)
        refs.update(superscript_refs)

        # 特殊符号
        if "*" in text:
            refs.add("*")
        if "※" in text:
            refs.add("※")

        return refs

    def _build_html_parts(
        self,
        sheet,
        filename: str,
        flattened_headers: dict[int, str],
        notes_dict: dict[str, str],
        header_note_refs: set[str],
        header_rows: int,
        data_end_row: int,
    ) -> list[str]:
        """构建 HTML 各部分"""
        html_parts: list[str] = []

        # 上下文信息
        context_html = (
            f'<div class="rag-context">【文档上下文】来源：{filename} | 数据类型：表格数据</div>'
        )
        html_parts.append(context_html)

        # 注释元数据
        if notes_dict:
            html_parts.append(self._build_notes_meta(notes_dict, header_note_refs))

        # 表格开始
        html_parts.append(
            f'<table border="1" style="border-collapse:collapse" '
            f'data-source="{filename}" data-sheet="{sheet.title}">'
        )

        # 关键词
        if self.keywords:
            keyword_str = "，".join(self.keywords)
            html_parts.append(f"    <caption>关键检索词：{keyword_str}</caption>")

        # 表头
        html_parts.extend(self._build_thead(flattened_headers))

        # 表体
        html_parts.extend(self._build_tbody(sheet, header_rows, data_end_row))

        html_parts.append("</table>")
        return html_parts

    def _build_notes_meta(self, notes_dict: dict[str, str], header_note_refs: set[str]) -> str:
        """构建注释元数据"""
        header_notes = {k: v for k, v in notes_dict.items() if k in header_note_refs}
        other_notes = {k: v for k, v in notes_dict.items() if k not in header_note_refs}

        notes_meta = {"header_notes": header_notes, "conditional_notes": other_notes}
        notes_json = json.dumps(notes_meta, ensure_ascii=False)
        return f'<script type="application/json" class="table-notes-meta">{notes_json}</script>'

    def _build_thead(self, flattened_headers: dict[int, str]) -> list[str]:
        """构建表头"""
        parts = ["    <thead>", "        <tr>"]
        for col_idx in sorted(flattened_headers.keys()):
            flat_header = flattened_headers.get(col_idx, "")
            parts.append(f"            <th>{flat_header}</th>")
        parts.extend(["        </tr>", "    </thead>"])
        return parts

    def _build_tbody(self, sheet, header_rows: int, data_end_row: int) -> list[str]:
        """构建表体"""
        parts = ["    <tbody>"]

        for row_idx in range(header_rows + 1, sheet.max_row + 1):
            is_note_row = row_idx > data_end_row
            row_class = ' class="table-note-row"' if is_note_row else ""
            parts.append(f"        <tr{row_class}>")

            for col_idx in range(1, sheet.max_column + 1):
                cell_html = self._build_cell(sheet, row_idx, col_idx)
                if cell_html:
                    parts.append(cell_html)

            parts.append("        </tr>")

        parts.append("    </tbody>")
        return parts

    def _build_cell(self, sheet, row_idx: int, col_idx: int) -> str | None:
        """构建单元格"""
        info = self._merged_info.get((row_idx, col_idx))

        if info and info.skip:
            return None

        if info:
            span_attrs = []
            if info.rowspan > 1:
                span_attrs.append(f'rowspan="{info.rowspan}"')
            if info.colspan > 1:
                span_attrs.append(f'colspan="{info.colspan}"')
            span_str = " " + " ".join(span_attrs) if span_attrs else ""
            cell_content = info.value if info.value is not None else ""
        else:
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_content = self._format_cell_value(cell)
            span_str = ""

        return f"            <td{span_str}>{cell_content}</td>"


def convert_excel_to_html(
    excel_path: str,
    keywords: list[str] | None = None,
    output_path: str | None = None,
) -> str | None:
    """将单个 Excel 文件转换为 RAG 增强的 HTML（兼容旧接口）"""
    converter = ExcelToHtmlConverter(keywords=keywords)
    result = converter.convert(Path(excel_path), Path(output_path) if output_path else None)
    return str(result) if result else None
