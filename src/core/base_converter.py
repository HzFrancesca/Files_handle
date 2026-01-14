"""
Excel 转换器抽象基类
提供所有格式转换器的共享逻辑
"""

import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from loguru import logger

from .models import MergedCellInfo


@dataclass
class BaseExcelConverter(ABC):
    """Excel 转换器抽象基类"""

    keywords: list[str] | None = None
    _merged_info: dict[tuple[int, int], MergedCellInfo] = field(
        default_factory=dict, init=False, repr=False
    )

    # ===== 模板方法 =====
    def convert(self, excel_path: Path, output_path: Path | None = None) -> Path | None:
        """执行转换（模板方法）"""
        source_path = Path(excel_path) if not isinstance(excel_path, Path) else excel_path

        if not source_path.exists():
            logger.error(f"找不到文件 '{source_path}'")
            return None

        out_path = self._determine_output_path(source_path, output_path)
        filename = source_path.name

        logger.info(f"正在处理: {filename}")
        self._log_features()

        try:
            workbook = openpyxl.load_workbook(str(source_path), data_only=False)
        except Exception as e:
            logger.error(f"解析失败: {e}")
            return None

        content = self._convert_workbook(workbook, filename)
        return self._write_output(out_path, content)

    def _determine_output_path(self, source_path: Path, output_path: Path | None) -> Path:
        """确定输出路径"""
        if output_path:
            return Path(output_path)
        ext = self._get_file_extension()
        return source_path.with_suffix("").with_name(f"{source_path.stem}_middle{ext}")

    def _log_features(self) -> None:
        """记录启用的功能"""
        features = "上下文硬编码 ✓ | 表头降维 ✓ | 合并单元格 ✓"
        if self.keywords:
            features += f" | 关键词 ✓ ({len(self.keywords)}个)"
        else:
            features += " | 关键词 ✗"
        logger.info(f"增强功能: {features}")

    def _convert_workbook(self, workbook, filename: str) -> str:
        """转换整个工作簿"""
        sheet_contents = []
        for sheet in workbook.worksheets:
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue
            self._merged_info = self._extract_merged_cells(sheet)
            header_rows = self._detect_header_rows(sheet)
            flattened_headers = self._build_flattened_headers(sheet, header_rows)
            footer_notes, data_end_row = self._detect_footer_notes(sheet, header_rows)
            
            content = self._format_sheet(
                sheet, filename, flattened_headers, header_rows, data_end_row
            )
            sheet_contents.append(content)
        
        return self._join_sheets(sheet_contents)

    def _write_output(self, out_path: Path, content: str) -> Path | None:
        """写入输出文件"""
        try:
            out_path.write_text(content, encoding="utf-8")
            logger.info(f"转换成功！输出: {out_path.absolute()}")
            return out_path
        except OSError as e:
            logger.error(f"写入文件失败: {e}")
            return None


    # ===== 共享实现 =====
    def _extract_merged_cells(self, sheet) -> dict[tuple[int, int], MergedCellInfo]:
        """获取所有合并单元格的信息"""
        merged_info: dict[tuple[int, int], MergedCellInfo] = {}

        for merged_range in sheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            origin_value = sheet.cell(row=min_row, column=min_col).value

            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    is_origin = r == min_row and c == min_col
                    merged_info[(r, c)] = MergedCellInfo(
                        value=str(origin_value) if origin_value is not None else None,
                        rowspan=rowspan if is_origin else 0,
                        colspan=colspan if is_origin else 0,
                        is_origin=is_origin,
                        skip=not is_origin,
                    )

        return merged_info

    def _detect_header_rows(self, sheet, max_check_rows: int = 5) -> int:
        """检测表头行数"""
        header_rows = 1

        for row_idx in range(1, min(max_check_rows + 1, sheet.max_row + 1)):
            has_colspan = False
            for col_idx in range(1, sheet.max_column + 1):
                info = self._merged_info.get((row_idx, col_idx))
                if info and info.colspan > 1:
                    has_colspan = True
                    break

            if has_colspan:
                header_rows = max(header_rows, row_idx + 1)

        return min(header_rows, sheet.max_row)

    def _build_flattened_headers(self, sheet, header_rows: int) -> dict[int, str]:
        """构建降维后的表头"""
        if header_rows <= 1:
            return self._build_single_row_headers(sheet)
        return self._build_multi_row_headers(sheet, header_rows)

    def _build_single_row_headers(self, sheet) -> dict[int, str]:
        """构建单行表头"""
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col_idx).value
            headers[col_idx] = str(value) if value else f"列{col_idx}"
        return headers

    def _build_multi_row_headers(self, sheet, header_rows: int) -> dict[int, str]:
        """构建多行表头（降维）"""
        col_values: dict[int, list[str]] = {col: [] for col in range(1, sheet.max_column + 1)}

        for row_idx in range(1, header_rows + 1):
            for col_idx in range(1, sheet.max_column + 1):
                info = self._merged_info.get((row_idx, col_idx))
                if info:
                    value = info.value
                    col_values[col_idx].append(str(value) if value else "")
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    col_values[col_idx].append(self._format_cell_value(cell))

        headers = {}
        for col_idx, values in col_values.items():
            unique_values = []
            for v in values:
                v = v.strip()
                if v and (not unique_values or v != unique_values[-1]):
                    unique_values.append(v)
            headers[col_idx] = "-".join(unique_values) if unique_values else f"列{col_idx}"

        return headers

    def _format_cell_value(self, cell) -> str:
        """格式化单元格值"""
        value = cell.value
        if value is None:
            return ""

        number_format = cell.number_format or "General"

        if isinstance(value, datetime):
            return self._format_datetime(value, number_format)

        if not isinstance(value, (int, float)):
            return str(value)

        return self._format_number(value, number_format)

    def _format_datetime(self, value: datetime, number_format: str) -> str:
        """格式化日期时间"""
        if "H" in number_format or "h" in number_format:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")

    def _format_number(self, value: int | float, number_format: str) -> str:
        """格式化数字"""
        if "%" in number_format:
            return self._format_percentage(value, number_format)

        if "E" in number_format.upper() and number_format != "General":
            return self._format_scientific(value, number_format)

        if "#,##" in number_format or ",0" in number_format:
            return self._format_currency(value, number_format)

        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)

    def _format_percentage(self, value: float, number_format: str) -> str:
        """格式化百分比"""
        decimal_match = re.search(r"0\.(0+)%", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        return f"{value * 100:.{decimals}f}%"

    def _format_scientific(self, value: float, number_format: str) -> str:
        """格式化科学计数法"""
        decimal_match = re.search(r"0\.(0+)E", number_format, re.IGNORECASE)
        decimals = len(decimal_match.group(1)) if decimal_match else 2
        return f"{value:.{decimals}E}"

    def _format_currency(self, value: float, number_format: str) -> str:
        """格式化货币"""
        decimal_match = re.search(r"0\.(0+)", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        formatted = f"{value:,.{decimals}f}"
        if "¥" in number_format or "￥" in number_format:
            return f"¥{formatted}"
        elif "$" in number_format:
            return f"${formatted}"
        return formatted

    def _detect_footer_notes(self, sheet, header_rows: int) -> tuple[list[str], int]:
        """检测表格末尾的注释行"""
        notes: list[str] = []
        note_patterns = [
            "注", "备注", "说明", "注意", "*", "※", "●", "◆", "△", "▲",
            "[注", "（注", "(注",
        ]

        for row_idx in range(sheet.max_row, header_rows, -1):
            content, is_note = self._check_note_row(sheet, row_idx, note_patterns)

            if not content:
                continue

            if is_note:
                notes.insert(0, content)
            else:
                break

        data_end_row = sheet.max_row - len(notes)
        return notes, data_end_row

    def _check_note_row(self, sheet, row_idx: int, note_patterns: list[str]) -> tuple[str, bool]:
        """检查是否为注释行"""
        filled_cols = 0
        content = ""
        is_merged_wide = False

        for col_idx in range(1, sheet.max_column + 1):
            info = self._merged_info.get((row_idx, col_idx))
            cell = sheet.cell(row=row_idx, column=col_idx)

            if info and info.is_origin and info.colspan > sheet.max_column // 2:
                is_merged_wide = True
                content = info.value if info.value else ""
                break

            if info and info.skip:
                continue

            if cell.value:
                filled_cols += 1
                if not content:
                    content = str(cell.value)

        content = content.strip()
        if not content:
            return "", False

        is_note = is_merged_wide or (
            filled_cols <= 2 and any(content.startswith(p) for p in note_patterns)
        )

        return content, is_note

    def _get_row_values(self, sheet, row_idx: int) -> list[str]:
        """获取一行的所有值（处理合并单元格）"""
        values = []
        for col_idx in range(1, sheet.max_column + 1):
            info = self._merged_info.get((row_idx, col_idx))
            if info:
                values.append(info.value or "")
            else:
                cell = sheet.cell(row=row_idx, column=col_idx)
                values.append(self._format_cell_value(cell))
        return values


    # ===== 抽象方法（子类实现）=====
    @abstractmethod
    def _get_file_extension(self) -> str:
        """返回输出文件扩展名"""
        ...

    @abstractmethod
    def _format_sheet(
        self,
        sheet,
        filename: str,
        flattened_headers: dict[int, str],
        header_rows: int,
        data_end_row: int,
    ) -> str:
        """格式化单个 sheet 为目标格式"""
        ...

    @abstractmethod
    def _join_sheets(self, sheet_contents: list[str]) -> str:
        """合并多个 sheet 的内容"""
        ...
