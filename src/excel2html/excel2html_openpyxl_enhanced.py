"""
Excel 转 HTML 表格 - RAG 增强版
支持文档增强功能，提升 RAG 检索召回率

增强功能：
1. 上下文硬编码 - 注入文件名、Sheet名等元数据
2. 幽灵标题 - 添加同义词和关键检索词
3. 表头降维 - 把父级标题拼接到子级标题（针对多层表头）
4. 合并单元格智能处理
"""

from pathlib import Path
from datetime import datetime
import re
import openpyxl


def format_cell_value(cell):
    """
    根据单元格的 number_format 返回格式化后的显示值
    解决 openpyxl 只返回存储值而非显示值的问题
    """
    value = cell.value
    if value is None:
        return ""

    number_format = cell.number_format or "General"

    # 日期/时间类型
    if isinstance(value, datetime):
        # 根据格式判断是否需要时间部分
        if "H" in number_format or "h" in number_format:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return value.strftime("%Y-%m-%d")

    # 非数字类型直接返回
    if not isinstance(value, (int, float)):
        return str(value)

    # 百分比格式
    if "%" in number_format:
        # 提取小数位数
        decimal_match = re.search(r"0\.(0+)%", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        return f"{value * 100:.{decimals}f}%"

    # 科学计数法 (必须明确指定 E 格式，且不是 General)
    if "E" in number_format.upper() and number_format != "General":
        decimal_match = re.search(r"0\.(0+)E", number_format, re.IGNORECASE)
        decimals = len(decimal_match.group(1)) if decimal_match else 2
        return f"{value:.{decimals}E}"

    # 货币和千分位格式
    if "#,##" in number_format or ",0" in number_format:
        # 检测小数位数
        decimal_match = re.search(r"0\.(0+)", number_format)
        decimals = len(decimal_match.group(1)) if decimal_match else 0

        # 格式化数字
        formatted = f"{value:,.{decimals}f}"

        # 添加货币符号
        if "¥" in number_format or "￥" in number_format:
            return f"¥{formatted}"
        elif "$" in number_format:
            return f"${formatted}"
        else:
            return formatted

    # 默认：普通数字
    # 如果是整数就不显示小数点
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def get_merged_cell_info(sheet):
    """
    获取所有合并单元格的信息
    返回: {(row, col): {'value': 值, 'rowspan': 行跨度, 'colspan': 列跨度, 'is_origin': 是否是左上角}}
    """
    merged_info = {}

    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        origin_value = sheet.cell(row=min_row, column=min_col).value

        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                is_origin = r == min_row and c == min_col
                merged_info[(r, c)] = {
                    "value": origin_value,
                    "rowspan": rowspan if is_origin else 0,
                    "colspan": colspan if is_origin else 0,
                    "is_origin": is_origin,
                    "skip": not is_origin,  # 非左上角的单元格需要跳过
                }

    return merged_info


def detect_header_rows(sheet, merged_info, max_check_rows=5):
    """
    检测表头行数（通过合并单元格和内容特征判断）
    返回表头行数
    """
    header_rows = 1

    for row_idx in range(1, min(max_check_rows + 1, sheet.max_row + 1)):
        has_colspan = False
        for col_idx in range(1, sheet.max_column + 1):
            info = merged_info.get((row_idx, col_idx))
            if info and info.get("colspan", 1) > 1:
                has_colspan = True
                break

        if has_colspan:
            header_rows = max(header_rows, row_idx + 1)

    return min(header_rows, sheet.max_row)


def detect_footer_notes(sheet, merged_info, header_rows):
    """
    检测表格末尾的注释行
    
    特征：
    - 只有1-2列有内容
    - 内容以"注"、"备注"、"说明"、"*"等开头
    - 或者是跨越多列的合并单元格
    
    返回: (notes_list, data_end_row)
    """
    notes = []
    # 扩展匹配模式，包括 [注]、（注）等带括号的形式
    note_patterns = ["注", "备注", "说明", "注意", "*", "※", "●", "◆", "△", "▲", "[注", "（注", "(注"]
    
    # 从末尾往前扫描
    for row_idx in range(sheet.max_row, header_rows, -1):
        filled_cols = 0
        content = ""
        is_merged_wide = False
        
        for col_idx in range(1, sheet.max_column + 1):
            info = merged_info.get((row_idx, col_idx))
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 检查是否是跨多列的合并单元格（超过一半列数）
            if info and info.get("is_origin") and info.get("colspan", 1) > sheet.max_column // 2:
                is_merged_wide = True
                content = str(info["value"]) if info["value"] else ""
                break
            
            # 跳过被合并的非原点单元格
            if info and info.get("skip"):
                continue
                
            if cell.value:
                filled_cols += 1
                if not content:  # 取第一个有内容的单元格
                    content = str(cell.value)
        
        content = content.strip()
        if not content:
            continue  # 空行跳过，继续往前找
            
        # 判断是否是注释行
        is_note = False
        if is_merged_wide:
            is_note = True
        elif filled_cols <= 2:
            # 检查是否以注释标识开头
            if any(content.startswith(p) for p in note_patterns):
                is_note = True
        
        if is_note:
            notes.insert(0, content)  # 保持原顺序
        else:
            break  # 遇到正常数据行就停止
    
    data_end_row = sheet.max_row - len(notes)
    return notes, data_end_row


def build_flattened_headers(sheet, merged_info, header_rows):
    """
    构建降维后的表头（把父级标题拼接到子级标题）
    返回: {col_idx: "父标题-子标题-..."}
    """
    if header_rows <= 1:
        # 单行表头，直接返回
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col_idx).value
            headers[col_idx] = str(value) if value else f"列{col_idx}"
        return headers

    # 多行表头，需要降维
    # 先构建每列在每行的实际值（考虑合并单元格）
    col_values = {col: [] for col in range(1, sheet.max_column + 1)}

    for row_idx in range(1, header_rows + 1):
        for col_idx in range(1, sheet.max_column + 1):
            info = merged_info.get((row_idx, col_idx))
            if info:
                value = info["value"]
                col_values[col_idx].append(str(value) if value else "")
            else:
                cell = sheet.cell(row=row_idx, column=col_idx)
                col_values[col_idx].append(format_cell_value(cell))

    # 拼接表头，去除重复和空值
    headers = {}
    for col_idx, values in col_values.items():
        # 去除空值和重复
        unique_values = []
        for v in values:
            v = v.strip()
            if v and (not unique_values or v != unique_values[-1]):
                unique_values.append(v)

        headers[col_idx] = "-".join(unique_values) if unique_values else f"列{col_idx}"

    return headers


def parse_notes_with_keys(notes_list):
    """
    解析注释列表，提取注释编号和内容
    
    例如：
    - "注1：xxx" -> {"注1": "注1：xxx"}
    - "注：xxx" -> {"注": "注：xxx"}
    - "[注] xxx" -> {"注": "[注] xxx"}
    - "*xxx" -> {"*": "*xxx"}
    
    返回: {key: full_note_text}
    """
    import re
    notes_dict = {}
    
    for note in notes_list:
        note = note.strip()
        if not note:
            continue
        
        # 匹配 "[注]"、"[注1]"、"（注）" 等带括号格式
        bracket_match = re.match(r'^[\[（\(](注\d*|备注\d*|说明\d*)[\]）\)]', note)
        if bracket_match:
            key = bracket_match.group(1)
            notes_dict[key] = note
            continue
            
        # 匹配 "注1："、"注："、"注1."、"注1、" 等格式
        match = re.match(r'^(注\d*|备注\d*|说明\d*|注意\d*)[：:．.、]?\s*', note)
        if match:
            key = re.match(r'^(注\d*|备注\d*|说明\d*|注意\d*)', note).group(1)
            notes_dict[key] = note
            continue
            
        # 匹配 "*"、"※"、"●" 等符号开头
        if note[0] in "*※●◆△▲":
            notes_dict[note[0]] = note
            continue
            
        # 无明确编号的注释，用内容前几个字作为 key
        notes_dict[note[:10]] = note
    
    return notes_dict


def extract_note_references(text):
    """
    从文本中提取注释引用
    
    例如：
    - "EX[注1]" -> ["注1"]
    - "税则号列对应关系[注]" -> ["注"]
    - "数据*" -> ["*"]
    
    返回: set of references
    """
    import re
    refs = set()
    
    # 匹配 [注1]、[注]、[备注1] 等
    bracket_refs = re.findall(r'\[(注\d*|备注\d*|说明\d*|注意\d*)\]', text)
    refs.update(bracket_refs)
    
    # 匹配上标形式或直接跟随的注释标记
    superscript_refs = re.findall(r'[^\[](注\d+)(?:[：:）\)]|$|\s)', text)
    refs.update(superscript_refs)
    
    # 匹配 * 等符号（通常表示有注释）
    if '*' in text or '※' in text:
        if '*' in text:
            refs.add('*')
        if '※' in text:
            refs.add('※')
    
    return refs


def sheet_to_enhanced_html(sheet, filename, keywords=None):
    """
    将单个 sheet 转换为 RAG 增强的 HTML 表格

    参数:
        sheet: openpyxl worksheet
        filename: 源文件名
        keywords: 可选的关键检索词列表
    """
    merged_info = get_merged_cell_info(sheet)
    header_rows = detect_header_rows(sheet, merged_info)
    flattened_headers = build_flattened_headers(sheet, merged_info, header_rows)
    
    # 检测末尾注释
    footer_notes, data_end_row = detect_footer_notes(sheet, merged_info, header_rows)
    
    # 解析注释为 key-value 格式
    notes_dict = parse_notes_with_keys(footer_notes)
    
    # 提取表头中的注释引用（这些注释所有 chunk 都需要）
    header_text = " ".join(flattened_headers.values())
    header_note_refs = extract_note_references(header_text)

    html_parts = []

    # === 增强1: 上下文硬编码 ===
    context_html = f"""<div class="rag-context">【文档上下文】来源：{filename} | 数据类型：表格数据</div>"""
    html_parts.append(context_html)
    
    # 注释元数据（供 chunk 切分时智能匹配使用）
    if notes_dict:
        import json
        # 表头引用的注释
        header_notes = {k: v for k, v in notes_dict.items() if k in header_note_refs}
        # 其他注释（需要按数据行匹配）
        other_notes = {k: v for k, v in notes_dict.items() if k not in header_note_refs}
        
        notes_meta = {
            "header_notes": header_notes,  # 所有 chunk 都添加
            "conditional_notes": other_notes  # 按匹配添加
        }
        notes_json = json.dumps(notes_meta, ensure_ascii=False)
        html_parts.append(f'<script type="application/json" class="table-notes-meta">{notes_json}</script>')

    # 开始表格
    html_parts.append(
        f'<table border="1" style="border-collapse:collapse" data-source="{filename}" data-sheet="{sheet.title}">'
    )

    # === 增强2: 幽灵标题 (Ghost Caption) ===
    if keywords:
        keyword_str = "，".join(keywords)
        caption_html = f"    <caption>关键检索词：{keyword_str}</caption>"
        html_parts.append(caption_html)

    # === 增强3: 表头降维 - 只保留扁平化的表头 ===
    html_parts.append("    <thead>")
    html_parts.append("        <tr>")
    for col_idx in range(1, sheet.max_column + 1):
        flat_header = flattened_headers.get(col_idx, "")
        html_parts.append(f"            <th>{flat_header}</th>")
    html_parts.append("        </tr>")
    html_parts.append("    </thead>")

    # 数据行（保留所有行，包括末尾注释行，供中间结果参照）
    html_parts.append("    <tbody>")
    for row_idx in range(header_rows + 1, sheet.max_row + 1):
        # 标记注释行，供 chunk 切分时识别
        is_note_row = row_idx > data_end_row
        row_class = ' class="table-note-row"' if is_note_row else ""
        html_parts.append(f"        <tr{row_class}>")
        for col_idx in range(1, sheet.max_column + 1):
            info = merged_info.get((row_idx, col_idx))

            if info and info.get("skip"):
                continue

            if info:
                value = info["value"]
                rowspan = info.get("rowspan", 1)
                colspan = info.get("colspan", 1)
                span_attrs = []
                if rowspan > 1:
                    span_attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    span_attrs.append(f'colspan="{colspan}"')
                span_str = " " + " ".join(span_attrs) if span_attrs else ""
                cell_content = str(value) if value is not None else ""
            else:
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_content = format_cell_value(cell)
                span_str = ""

            html_parts.append(f"            <td{span_str}>{cell_content}</td>")
        html_parts.append("        </tr>")

    html_parts.append("    </tbody>")
    html_parts.append("</table>")

    return "\n".join(html_parts)


def convert_excel_to_html(
    excel_path: str, keywords: list = None, output_path: str = None
):
    """
    将单个 Excel 文件转换为 RAG 增强的 HTML

    参数:
        excel_path: Excel 文件路径
        keywords: 可选的关键检索词列表，用于幽灵标题
        output_path: 可选的输出路径，默认与源文件同目录同名

    返回:
        成功返回输出文件路径，失败返回 None
    """
    source_path = Path(excel_path)

    if not source_path.exists():
        print(f"❌ 错误：找不到文件 '{source_path}'")
        return None

    if output_path:
        out_path = Path(output_path)
    else:
        out_path = source_path.with_suffix("").with_name(
            source_path.stem + "_converted.html"
        )

    filename = source_path.name
    print(f"📄 正在处理: {filename}")
    print(f"   增强功能: 上下文硬编码 ✓ | 表头降维 ✓ | 合并单元格 ✓ | 注释提取 ✓", end="")
    if keywords:
        print(f" | 幽灵标题 ✓ ({len(keywords)}个关键词)")
    else:
        print(" | 幽灵标题 ✗ (未提供关键词)")

    try:
        workbook = openpyxl.load_workbook(str(source_path), data_only=False)
    except Exception as e:
        print(f"❌ 解析失败: {e}")
        return None

    # 构建 HTML - 只输出核心内容，不包含文档外壳
    html_parts = []

    for sheet in workbook.worksheets:
        if sheet.max_row == 0 or sheet.max_column == 0:
            continue  # 跳过空 sheet

        html_parts.append(sheet_to_enhanced_html(sheet, filename, keywords))

    # 写入文件
    try:
        out_path.write_text("\n".join(html_parts), encoding="utf-8")
        print(f"✅ 转换成功！输出: {out_path.absolute()}")
        return str(out_path)
    except IOError as e:
        print(f"❌ 写入文件失败: {e}")
        return None


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(
        description="Excel 转 HTML (RAG 增强版)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python excel2html_openpyxl_enhanced.py input.xlsx
  python excel2html_openpyxl_enhanced.py input.xlsx -o output.html
  python excel2html_openpyxl_enhanced.py input.xlsx -k "财务报表" "年度收入" "利润"
        """,
    )
    parser.add_argument("excel_file", help="要转换的 Excel 文件路径")
    parser.add_argument("-o", "--output", help="输出 HTML 文件路径（可选）")
    parser.add_argument(
        "-k", "--keywords", nargs="+", help="关键检索词（用于幽灵标题）"
    )

    args = parser.parse_args()

    convert_excel_to_html(
        excel_path=args.excel_file, keywords=args.keywords, output_path=args.output
    )


if __name__ == "__main__":
    main()
