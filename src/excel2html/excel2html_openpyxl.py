"""
Excel 转 HTML 表格。
支持两种模式：填充合并单元格 / 不填充（使用 unstructured 解析）
"""

from pathlib import Path
import openpyxl


def get_merged_cell_value(sheet, row, col):
    """获取单元格的值，如果是合并单元格则返回合并区域左上角的值"""
    cell = sheet.cell(row=row, column=col)

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value

    return cell.value


def get_cell_value(sheet, row, col, fill_merged=True):
    """
    获取单元格值
    fill_merged=True: 合并单元格填充相同内容
    fill_merged=False: 只有左上角有值，其他位置为空
    """
    if fill_merged:
        return get_merged_cell_value(sheet, row, col)
    else:
        return sheet.cell(row=row, column=col).value


def sheet_to_html_table(sheet, fill_merged=True):
    """将单个 sheet 转换为 HTML 表格"""
    html = [f"<h2>{sheet.title}</h2>", "<table>"]

    for row_idx in range(1, sheet.max_row + 1):
        html.append("  <tr>")
        for col_idx in range(1, sheet.max_column + 1):
            value = get_cell_value(sheet, row_idx, col_idx, fill_merged)
            cell_content = str(value) if value is not None else ""
            html.append(f"    <td>{cell_content}</td>")
        html.append("  </tr>")

    html.append("</table>")
    return "\n".join(html)


def convert_excel_to_html(excel_path: Path, fill_merged=True):
    """将 Excel 文件转换为 HTML"""
    source_path = excel_path if isinstance(excel_path, Path) else Path(excel_path)

    if not source_path.exists():
        print(f"❌ 错误：找不到文件 '{source_path}'")
        return

    output_path = source_path.with_suffix(".html")
    mode_text = "填充" if fill_merged else "不填充"
    print(f"正在处理 ({mode_text}模式): {source_path.name}")

    try:
        workbook = openpyxl.load_workbook(str(source_path), data_only=True)
    except Exception as e:
        print(f"❌ 解析失败: {e}")
        return

    html_content = [
        """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>Excel Export</title>
</head>
<body>"""
    ]

    for sheet in workbook.worksheets:
        html_content.append(sheet_to_html_table(sheet, fill_merged))

    html_content.append("</body></html>")

    try:
        output_path.write_text("\n".join(html_content), encoding="utf-8")
        print(f"✅ 转换成功！文件已保存至: {output_path.absolute()}")
    except IOError as e:
        print(f"❌ 写入文件失败: {e}")


def convert_folder(folder_path_str: str, fill_merged=True):
    """批量处理指定文件夹下的所有 Excel 文件"""
    folder = Path(folder_path_str)

    if not folder.exists():
        print(f"❌ 错误：找不到文件夹 '{folder}'")
        return

    if not folder.is_dir():
        print(f"❌ 错误：'{folder}' 不是一个文件夹")
        return

    excel_files = [
        f
        for f in list(folder.glob("*.xlsx")) + list(folder.glob("*.xls"))
        if not f.name.startswith("~$")
    ]

    if not excel_files:
        print(f"⚠️ 文件夹 '{folder}' 中没有找到 Excel 文件")
        return

    print(f"📁 找到 {len(excel_files)} 个 Excel 文件\n")

    for excel_file in excel_files:
        convert_excel_to_html(excel_file, fill_merged)

    print("\n🎉 处理完成！")


if __name__ == "__main__":
    target_folder = (
        r"C:\Users\Administrator\Desktop\玄通\通用知识库_handled\2026年税则调整"
    )

    # fill_merged=True  填充合并单元格（推荐用于 LLM）
    # fill_merged=False 不填充，合并区域只有左上角有值
    convert_folder(target_folder, fill_merged=True)
